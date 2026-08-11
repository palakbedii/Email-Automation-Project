from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .config import (
    REPORT_FOLDER,
    REPORT_NAME_PREFIX,
    SHEET_NAME
)


def generate_report(
    node_name,
    ip_address,
    cpu_usage,
    memory_usage,
    disk_usage
):
    """
    Generate an Excel monitoring report.

    Parameters:
        node_name (str): Name of the monitored node.
        ip_address (str): IP address of the monitored node.
        cpu_usage (float): CPU usage percentage.
        memory_usage (float): RAM usage percentage.
        disk_usage (float): Disk usage percentage.

    Returns:
        Path: Path of the generated Excel report.
    """

    # 1. Make sure the Reports folder exists

    REPORT_FOLDER.mkdir(parents=True, exist_ok=True)


    # 2. Create workbook

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = SHEET_NAME


    # 3. Define headers

    headers = [
        "NODE NAME",
        "IP",
        "CPU %",
        "RAM %",
        "DISK %"
    ]

    worksheet.append(headers)


    # 4. Add monitoring data

    worksheet.append([
        node_name,
        ip_address,
        cpu_usage,
        memory_usage,
        disk_usage
    ])


    # 5. Format headers

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


    # 6. Format data cells

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border


    # 7. Set column widths

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 12


    # 8. Create unique report filename

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = (
        f"{REPORT_NAME_PREFIX}_{timestamp}.xlsx"
    )

    report_path = REPORT_FOLDER / filename


    # 9. Save workbook

    workbook.save(report_path)

    return report_path